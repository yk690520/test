'''
需求说明：导入一个excel表格，表格表头：【资产编号，标签】，然后将其设置到后台，并输出未成功设置的资产编号。
'''
import openpyxl,time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from other_duty.decorator.decorator import decorator_for_log_and_load_file

@decorator_for_log_and_load_file(data_column=2,log_column_name=["资产编号","处理结果"])
def ding_flags(sourcePath,logPath,*,data_list):
    will_ding_assert_list=data_list
    address_list=["http://console.zy.youkedao.com/login/","http://console.yk.yijuchelian.com/login/"]
    for i in range(0,len(address_list)):
        print("【%s】 %s" % (i+1,address_list[i]))
    address=input("请选择或者直接输入后台网址：")
    if address=="1":
        address=address_list[0]
    elif address=="2":
        address=address_list[1]
    elif "http" not in address:
        print("输入网址不正确")
        raise BaseException("输入网址不正确")
    usernme=input("请输入账号：")
    pwd=input("请输入密码：")
    #进入浏览器
    brower = webdriver.Chrome()
    brower.maximize_window()
    brower.implicitly_wait(1)
    brower.get(address)
    brower.find_element_by_xpath('//*[@id="userName"]').send_keys(usernme)
    brower.find_element_by_xpath('//*[@id="password"]').send_keys(pwd)
    brower.find_element_by_xpath('/html/body/form/div/div[2]/div[4]/input').click()
    brower.find_element_by_xpath('//*[@id="side-menu"]/dl[7]/dd[1]/a').click()
    #抵达微信号管理页面
    access=[]#成功的
    unaccess=[]#失败的
    log=[]
    for one in will_ding_assert_list:
        asset=one[0]
        flags=one[1]
        flags_str = flags.strip()
        flags = flags_str.split(",")
        try:
            __ding_flags_for_wechat(brower,asset,flags)
        except:
            log.append([asset,"失败"])
        else:
            log.append([asset, "成功"])
    brower.quit()
    #打标签结束，返回错误日志
    return log

def __ding_flags_for_wechat(brower,asset:str,flags:[]):
    '''
    这是一个为某一个具体的设备订标签的方法，如果传入的标签为None，那么将清除此设备的所有标签
    :param asset: 设备资产编号
    :param flags: 要打的标签
    :return:None
    '''
    brower.find_element_by_xpath('//*[@id="compose_keywords"]').clear()
    brower.find_element_by_xpath('//*[@id="compose_keywords"]').send_keys(asset)
    brower.find_element_by_xpath('//*[@id="btn-filter"]/span[2]').click()
    #找到change按钮
    time.sleep(1)
    change=brower.find_element_by_xpath('//i[@onclick="modifyGroup($(this))"]')
    #点击change按钮
    brower.execute_script("$(arguments[0]).click()", change)
    #点击清除按钮
    brower.find_element_by_xpath('//*[@id="group-btn-group-cancel"]/span[1]').click()
    #点击ok按钮
    brower.find_element_by_xpath('//*[@id="group-btn-group-ok"]').click()
    brower.find_element_by_xpath('//*[@id="popup_ok"]').click()
    if not flags:
        return
    # 找到change按钮
    change = brower.find_element_by_xpath('//i[@onclick="modifyGroup($(this))"]')
    # 点击change按钮
    brower.execute_script("$(arguments[0]).click()", change)
    for flag in flags:
        __ding_flag(brower,flag)
        #brower.find_element_by_xpath('//span[@class="splitTags"][text()="%s"]' % flag).click()
    time.sleep(1)
    brower.find_element_by_xpath('//*[@id="group-btn-group-ok"]').click()  # 点击ok按钮

def __ding_flag(brower,flag):
    brower.find_element_by_xpath('//*[@id="search_group"]').clear()
    brower.find_element_by_xpath('//*[@id="search_group"]').send_keys(flag)
    brower.find_element_by_xpath('//*[@id="btn_search"]').click()
    __add_flag(brower,"自动标签",flag)
    brower.find_element_by_xpath('//span[@class="splitTags"][text()="%s"]' % flag).click()

def __add_flag(brower:webdriver.Chrome,group,flag):
    '''
    增加一个标签
    入口条件网页进入到标签编辑状态，出口情况,位于标签编辑状态
    :param brower:
    :param group:标签所属组
    :param flag:标签名
    :return:
    '''
    try:
        WebDriverWait(brower, 1).until(EC.presence_of_element_located((By.XPATH, '//span[@class="splitTags"][text()="%s"]' % flag)))
    except:
        brower.find_element_by_xpath('//*[@id="search_group"]').clear()
        brower.find_element_by_xpath('//*[@id="btn_search"]').click()
        brower.find_element_by_xpath('//a[@class="edit"]').click()
        brower.find_element_by_xpath('//span[@class="createfGroupType splitTags"]//a').click()
        brower.find_element_by_xpath('//input[@class="layui-layer-input"]').send_keys(group)
        brower.find_element_by_xpath('//a[@class="layui-layer-btn0"]').click()
        try:
            brower.find_element_by_xpath('//*[@id="popup_ok"]').click()
        except:
            print("")
        brower.find_element_by_xpath('//td[@class="grouptype"][text()="%s"]/..//a[text()="+"]' % group).click()
        brower.find_element_by_xpath('//input[@class="layui-layer-input"]').send_keys(flag)
        brower.find_element_by_xpath('//a[@class="layui-layer-btn0"]').click()
        brower.find_element_by_xpath('//a[@class="cancel"]').click()
        brower.find_element_by_xpath('//*[@id="search_group"]').clear()
        brower.find_element_by_xpath('//*[@id="search_group"]').send_keys(flag)
        brower.find_element_by_xpath('//*[@id="btn_search"]').click()
def out_demo_file(path):

    '''
    输出一个demo文件
    :param path:
    :return:
    '''
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1,1,"查询字段")
    sheet.cell(1,2,"标签")
    sheet.cell(2,1,"SH-A-02")
    sheet.cell(2,2,"封号,故障")
    wb.save(path)

if __name__=="__main__":
    try:
        ding_flags(r'C:\Users\yk690\OneDrive\桌面\新建 Microsoft Excel 工作表.xlsx',r'C:\Users\yk690\OneDrive\桌面\log.xlsx')
    except BaseException as e:
        print(e)
        time.sleep(10)
