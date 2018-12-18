#此模块用来创建微信号分组
from selenium import webdriver
import openpyxl,os,time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from other_duty.decorator.decorator import decorator_for_log_and_load_file

def __add_flag(brower:webdriver.Chrome,group,flag):
    '''
    增加一个标签
    入口条件网页进入到标签编辑状态，出口情况,位于标签编辑状态
    :param brower:
    :param group:标签所属组
    :param flag:标签名
    :return:
    '''
    try:
        WebDriverWait(brower, 2).until(EC.presence_of_element_located((By.XPATH, '//span[@class="splitTags"][text()="%s"]' % flag)))
    except:
        try:
            brower.find_element_by_xpath('//td[@class="grouptype"][text()="%s"]' % group)
        except:
            # 增加group流程
            brower.find_element_by_xpath('//span[@class="createfGroupType splitTags"]//a').click()
            brower.find_element_by_xpath('//input[@class="layui-layer-input"]').send_keys(group)
            brower.find_element_by_xpath('//a[@class="layui-layer-btn0"]').click()
            pass
        brower.find_element_by_xpath('//td[@class="grouptype"][text()="%s"]/..//a[text()="+"]' % group).click()
        brower.find_element_by_xpath('//input[@class="layui-layer-input"]').send_keys(flag)
        brower.find_element_by_xpath('//a[@class="layui-layer-btn0"]').click()

@decorator_for_log_and_load_file(data_column=2,log_column_name=["标签","处理结果"])
def add_flags(sourcePath,logPath,*,data_list):
    group_and_flag_list=data_list
    #进入浏览器
    address_list = ["http://console.zy.youkedao.com/login/", "http://console.yk.yijuchelian.com/login/"]
    for i in range(0, len(address_list)):
        print("【%s】 %s" % (i + 1, address_list[i]))
    address = input("请选择或者直接输入后台网址：")
    if address == "1":
        address = address_list[0]
    elif address == "2":
        address = address_list[1]
    elif "http" not in address:
        print("输入网址不正确")
        raise BaseException("输入网址不正确")
    usernme = input("请输入账号：")
    pwd = input("请输入密码：")
    brower = webdriver.Chrome()
    brower.maximize_window()
    brower.implicitly_wait(2)
    brower.get(address)
    brower.find_element_by_xpath('//*[@id="userName"]').send_keys(usernme)
    brower.find_element_by_xpath('//*[@id="password"]').send_keys(pwd)
    brower.find_element_by_xpath('/html/body/form/div/div[2]/div[4]/input').click()
    brower.find_element_by_xpath('//*[@id="side-menu"]/dl[7]/dd[1]/a').click()
    time.sleep(3)
    change=brower.find_element_by_xpath('//div[@class="splitGroups"]//i')
    brower.execute_script("$(arguments[0]).click()", change)
    brower.find_element_by_xpath('//a[@onclick="editGroup()"]').click()
    #抵达微信号标签管理页面
    log=[]
    for one in group_and_flag_list:
        group=one[0]
        flag=one[1]
        try:
            __add_flag(brower,group,flag)
        except BaseException as e:
            log.append(["%s-%s" % (group,flag),"失败"])
        else:
            log.append(["%s-%s" % (group, flag), "成功"])
    brower.quit()
    #增加标签结束，返回错误日志
    return log
    pass

def out_demo_file(path):

    '''
    输出一个demo文件
    :param path:
    :return:
    '''
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1,1,"标签分组")
    sheet.cell(1,2,"标签")
    sheet.cell(2,1,"社群分组")
    sheet.cell(2,2,"社群组YCA")
    wb.save(path)

if __name__ == '__main__':
    add_flags("%s/demo.xlsx" % os.getcwd(),"%s/log.xlsx" % os.getcwd())