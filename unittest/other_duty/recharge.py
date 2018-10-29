# -*- encoding=utf8 -*-
#此模块用来控制充值话费，使用前需先将设备至于充值页面
from poco.drivers.android.uiautomation import AndroidUiautomationPoco
from poco.pocofw import Poco
from airtest.core.api import *
from airtest.core.android.adb import *
from airtest.core.android.constant import (DEFAULT_ADB_PATH, IP_PATTERN,
                                           SDK_VERISON_NEW)
import openpyxl,os,shutil
import requests,json
from other_duty.decorator.decorator import decorator_for_log_and_load_file
def __error_re(poco:Poco):
    '''
    异常恢复方法
    :param poco:
    :return:
    '''
    home()#回到主页
    stop_app("com.tencent.mm")#停止微信
    start_app("com.tencent.mm")#打开微信
    poco("com.tencent.mm:id/cdj",text='发现').click()
    poco(text="朋友圈").click()
    poco("com.tencent.mm:id/lo").click()
    poco("com.tencent.mm:id/lo").click()
    poco(text="发消息").click()
    poco("com.tencent.mm:id/hi").click()
    poco(text="清空聊天记录").click()
    poco("com.tencent.mm:id/an3").click()
    poco("com.tencent.mm:id/ht").click()
    poco("com.tencent.mm:id/ac9").set_text("http://www.xingjunwl.com/?acode=8d4ffc2c127933adad71af12cd5de3fd")
    poco("com.tencent.mm:id/ace").click()#发送
    poco("com.tencent.mm:id/ki").click()#进入到充值界面
    poco(name="代理申请").wait_for_appearance()

def out_demo_file(path):

    '''
    输出一个demo文件
    :param path:
    :return:
    '''
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1,1,"电话号码")
    sheet.cell(2,1,"13094528708")
    wb.save(path)

def __recharge_one_phone(poco:Poco,tel_phone:str):
    '''
    单个号码充值方法
    :param tel_phone:
    :return: [卡号,剩余流量,当前状态,流量有效期]
    '''
    poco("android.widget.EditText").wait_for_appearance(timeout=60)
    poco("android.widget.EditText").set_text(tel_phone)
    poco("查询充值").click()
    #判断是否需要充值
    gprs = poco("com.tencent.mm:id/b0c").child("android.widget.FrameLayout").child("android.widget.FrameLayout").child(
        "android.webkit.WebView").child("android.webkit.WebView").children()[9]
    if float(gprs.get_name()[:-2])>1024:
        return
    #判断是否需要充值
    poco("蚂蚁10G有效期30天").click()
    if poco("立即充值").exists():
        poco("立即充值").focus([0.5,1]).click()
    if poco("立即充值").exists():
        poco("立即充值").click()
    poco("android.widget.RadioButton").click()
    poco("android.widget.ListView").child("android.view.View")[1].child("android.widget.RadioButton").click()

    poco("确认支付").click()
    sleep(5)
    if poco("该流量包卡密不足").exists():
        raise BaseException("该流量包卡密不足")
    poco(text="请输入支付密码").wait_for_appearance(timeout=10)
    #输入密码
    # 1
    with poco.freeze() as freeze_poco:
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[0].focus(
            [0.2, 0.5]).click()
        # 1
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[0].focus(
            [0.2, 0.5]).click()
        # 5
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[1].focus(
            [0.5, 0.5]).click()
        # 5
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[1].focus(
            [0.5, 0.5]).click()
        # 5
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[1].focus(
            [0.5, 0.5]).click()
        # 5
        freeze_poco("com.tencent.mm:id/aam").child("android.widget.LinearLayout").child("android.widget.LinearLayout")[1].focus(
            [0.5, 0.5]).click()
        time.sleep(2)
        #可能需要输入验证码
        if poco(text="手机短信验证").exists():
            count=1
            while True:
                rsponse=requests.get("http://www.junx.ink/get_code.html")
                resault=json.loads(rsponse.text)
                if not resault["state"]=="success":
                    count+=1
                    time.sleep(5)
                    if count>=21:
                        raise BaseException("输入验证码超时")
                    continue
                code=resault["code"]
                poco("com.tencent.mm:id/by").set_text(code)
                poco("com.tencent.mm:id/xc").click()
                break
        #
    poco("com.tencent.mm:id/dxw").wait_for_appearance(timeout=10)
    poco("com.tencent.mm:id/dxw").click()
    #充值结束

def __get_infor(poco,tel_phone:str):
    '''
    记录充值结果,如果记录充值结果出错，则返回一个充值成功，但是记录出错的日志
    :param tel_phone:
    :return:
    '''
    try:
        time.sleep(2)
        poco("com.tencent.mm:id/ht").click()
        poco("com.tencent.mm:id/ki").click()  # 进入到充值界面
        poco(name="代理申请").wait_for_appearance(timeout=10)
        poco("android.widget.EditText").set_text(tel_phone)
        poco("查询充值").click()
        poco("立即充值").drag_to(poco("充值记录"))
        phone = poco("android.widget.GridView").child("android.view.View")[0].child(type="android.view.View")[1].get_name()
        gprs = poco("android.widget.GridView").child("android.view.View")[3].child(type="android.view.View")[1].get_name()
        status = poco("android.widget.GridView").child("android.view.View")[4].child(type="android.view.View")[
            1].get_name()
        up_date = poco("android.widget.GridView").child("android.view.View")[5].child(type="android.view.View")[
            1].get_name()
        poco("立即充值").drag_to(poco("流量有效期"))
    except:
        return tel_phone,"已充值","未记录到结果","请手动记录"
    else:
        return phone,gprs,status,up_date

@decorator_for_log_and_load_file(data_column=1,log_column_name=["手机号","剩余流量",'卡状态','有效期'])
def recharge(source_path:str,log_path:str,*,data_list):
    '''
    完整的充值方法
    :param source_path:
    :param log_path:
    :return:
    '''

    poco = AndroidUiautomationPoco(use_airtest_input=True, screenshot_each_action=False)
    phone_list=data_list
    log=[]
    __error_re(poco)
    for phone in phone_list:
        try:
            __recharge_one_phone(poco,phone)
            resault=__get_infor(poco,phone)
            log.append(resault)
        except BaseException as e:
            if (e.args) and (e.args[0]=="该流量包卡密不足"):
                one_log=[phone,"该流量包","卡密","不足"]
            elif(e.args) and (e.args[0]=="输入验证码超时"):
                one_log=[phone,"输入","验证码","超时"]
            else:
                one_log=[phone,"充值异常","请检查是否","充值成功"]
            if (e.args) and (e.args[0] == "输入验证码超时"):
                log.append(one_log)
                break
            log.append(one_log)
            try:
                __error_re(poco)
            except:
                print("错误重置出错")
                break
    return log