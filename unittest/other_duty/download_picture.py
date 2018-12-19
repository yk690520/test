import openpyxl,time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from other_duty.decorator.decorator import decorator_for_log_and_load_file
import urllib.request as request
import os,datetime

def out_demo_file(path):

    '''
    输出一个demo文件
    :param path:
    :return:
    '''
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1,1,"保存名")
    sheet.cell(1,2,"url")
    sheet.cell(2,1,"1")
    sheet.cell(2,2,"http://www.baidu.com")
    wb.save(path)

#分配二维码
@decorator_for_log_and_load_file(data_column=2,log_column_name=["链接","状态"])
def download_pic(source_path:str,log_path:str,save_path,*,data_list):
    '''
    二维码分配方法
    :param source_path:
    :param log_path:
    :return:
    '''
    distr_list=data_list
    log=[]
    timeFile=datetime.datetime.now().strftime("%Y-%m-%d %H%M%S")
    if not os.path.exists("%s\\%s" %(save_path,timeFile)):
        os.mkdir("%s\\%s" %(save_path,timeFile))
    save_path="%s\\%s" %(save_path,timeFile)
    for row in distr_list:
        fileName = row[0]
        url = row[1]
        try:
            request.urlretrieve(url,"%s\\%s.jpg" % (save_path, fileName))
            resault=[url,"成功"]
        except:
            resault=[url,"失败"]
        log.append(resault)
    return log