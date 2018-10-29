import functools
import openpyxl
def decorator_for_log_and_load_file(data_column:int,log_column_name:[],skip_first_row:bool=True):
    '''
    此装饰器所修饰的函数的参数必须含有source_path以及log_path的申明
    此装饰器将会读取source_path所指定的文件，并将其数据返回，如果数据列只有1列，将返回一维数组，2列及以上，返回二维数组
    :param data_column:
    :param skip_first_row:
    :return:
    '''
    def decorator(fux):
        @functools.wraps(fux)
        def wrapper(*args,**kwargs):
            #方法执行前
            excel=None
            source_path=args[0]
            log_path=args[1]
            try:
                excel = openpyxl.load_workbook(source_path, read_only=True)
                sheet = excel.active
                data_list = []
                skip_first = skip_first_row
                for row in sheet.rows:
                    if skip_first:
                        skip_first = False
                        continue
                    if data_column<1:
                        raise BaseException("数据列设置错误")
                    if data_column==1:
                        if not row[0].value:
                            break
                        data_list.append(row[0].value)
                    if data_column>1:
                        if not row[0].value:
                            break
                        data_row=[]
                        for i in range(data_column):
                            if row[i].value:
                                data_row.append(row[i].value)
                            else:
                                data_row.append("")
                        data_list.append(data_row)
            except BaseException as e:
                raise e
            finally:
                if excel:
                    excel.close()
            re=fux(*args,**kwargs,data_list=data_list)
            try:
                wb = openpyxl.Workbook()
                sheet = wb.active
                for i in range(len(log_column_name)):
                    sheet.cell(1,i+1,log_column_name[i])
                for i in range(0, len(re)):
                    for j in range(0, len(re[i])):
                        sheet.cell(i + 2, j + 1, re[i][j])
                wb.save(log_path)
            except BaseException as e:
                raise e
            return re
        return wrapper
    return decorator